from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException,StaleElementReferenceException
from selenium.webdriver.common.keys import Keys


import openpyxl
# load excel with its path
wrkbk = openpyxl.load_workbook("E:\Web scraping\zomato.xlsx")
# to get the active work sheet
sh = wrkbk.active

PATH="E:\chromedriver.exe"

driver=webdriver.Chrome(PATH)
driver.get('https://www.zomato.com/udaipur/restaurants?place_name=Udaipur&dishv2_id=7cf2db5ec261a0fa27a502d3196a6f60_2')
print(driver.title)

ignored_exceptions = (NoSuchElementException, StaleElementReferenceException,)
row = 1
for i in range(5):

    WebDriverWait(driver, 3, ignored_exceptions=ignored_exceptions).until(EC.presence_of_element_located((By.CLASS_NAME, 'sc-1hp8d8a-0')))
    restaurants = driver.find_elements_by_class_name('sc-1hp8d8a-0')
    restaurants[i].click()


    restaurant_name = WebDriverWait(driver, 5, ignored_exceptions=ignored_exceptions).until(
        EC.presence_of_element_located((By.CLASS_NAME, "ckvGKr"))
    )

    reviews = driver.find_elements_by_class_name('kEgyiI')
    timings = driver.find_element_by_class_name('sc-dEoRIm')
    address = driver.find_element_by_class_name('sc-iqzUVk')
    dining_review = reviews[0]
    delivery_review = reviews[1]
    # print('Restaurant Name', restaurant_name.text)
    # print('dining_review ', dining_review.text)
    # print('delivery_review ', delivery_review.text)
    # print('timings ',timings.text)
    # print('address', address.text)
    restaurant_main_details=[restaurant_name,timings,dining_review,delivery_review,address]
    for col,element in enumerate(restaurant_main_details):
        sh.cell(row=row , column=col+1 ).value = element.text
    col_default=len(restaurant_main_details)+1
    WebDriverWait(driver, 5, ignored_exceptions=ignored_exceptions).until(
        EC.presence_of_element_located((By.CLASS_NAME, "sc-1s0saks-10"))
    )

    try:
        read_more = driver.find_elements_by_class_name('sc-ya2zuu-0 ')
        for reads in read_more:
            reads.click()
    except NoSuchElementException:
        print("No read more")

    sections=driver.find_elements_by_class_name('sc-GLkNx')

    for category in sections:

        category_name=category.find_element_by_class_name('sc-1hp8d8a-0')
        # print('category ', category_name.text)
        item_cards=category.find_elements_by_class_name('sc-1s0saks-10')

        for item_card in item_cards:
            col = col_default

            try:
                chef_special = item_card.find_element_by_class_name('fQRUpA')
                # print(chef_special.text)
            except NoSuchElementException:
                chef_special=None

            try:
                must_try = item_card.find_element_by_class_name('cRxPpO')
                # print(must_try.text)
            except NoSuchElementException:
                must_try=None

            item_name=item_card.find_element_by_class_name('sc-1s0saks-15')
            item_description=item_card.find_element_by_class_name('sc-1s0saks-12')
            item_price = item_card.find_element_by_class_name('sc-17hyc2s-1')

            # print(item_name.text)
            # print(item_description.text)
            # print(item_price.text)
            for element in [category_name,item_name,chef_special,must_try,item_price,item_description]:
                if element!=None:
                    sh.cell(row=row, column=col).value = element.text
                col=col+1
            row=row+1
        # print()
        wrkbk.save('zomato.xlsx')
    driver.back()
    # driver.implicitly_wait(3)


wrkbk.save('zomato.xlsx')

# print(reviews.text)
# driver.quit()